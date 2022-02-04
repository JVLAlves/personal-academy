import os
import re

import exConsts
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import QMessageBox
from openpyxl.styles import numbers
"""
def InsertPatient(name, value, workbook):
    sheet = workbook.active
    sheet.insert_rows(exConsts.BEGINNING_ROW)
    Patref = exConsts.PATIENT_COLUMN + str(exConsts.PAT_ROW)
    Valref = exConsts.PAYPERSESSION_COLUMN + str(exConsts.PAT_ROW)
    Totref = exConsts.TOTALPAYMENT_COLUMN + str(exConsts.PAT_ROW)
    sheet[Patref] = name
    sheet[Valref] = value
    sheet[Totref] = '=SUM(C{}:N{})'.format(exConsts.PAT_ROW, exConsts.PAT_ROW)

    workbook.save(exConsts.FILENAME)
"""

def InsertPatient(name, value, filepath):
    workbook = load_workbook(filepath)
    sheet = workbook.active
    row = exConsts.BEGINNING_ROW
    cel = exConsts.PATIENT_COLUMN + str(row)
    while sheet[cel].value is not None:
        row+= 1
        cel = exConsts.PATIENT_COLUMN + str(row)

    nameCell = cel
    PayPerSessionCell = exConsts.PAYPERSESSION_COLUMN + str(row)
    TotalPaymentCell = exConsts.TOTALPAYMENT_COLUMN + str(row)
    TotalPaymentFormula = '=SUM(C{}:N{})'.format(row, row)

    sheet[nameCell] = name

    sheet[PayPerSessionCell] = value
    sheet[PayPerSessionCell].number_format = 'currency'
    sheet[TotalPaymentCell] = TotalPaymentFormula
    sheet[TotalPaymentCell].number_format = 'currency'

    workbook.save(filepath)

def SearchPatientByName(patname, workbook):
    #Variaveis para Uso interno
    sheet = workbook.active #Planilha
    PatCol = exConsts.PATIENT_COLUMN #Letra da Coluna do Nome dos Pacientes
    PatRow = exConsts.PAT_ROW #Número da linha dos Pacientes
    cel = PatCol + str(PatRow) #Célula determinada com número e letra
    possibleCels = [] #Possiveis células que combinem com a pesquisa.

    # Busca a célula na planilha com o mesmo nome.
    while sheet[cel].value is not None:
        if sheet[cel].value.startswith(patname.capitalize()):
            cell = {
                "cll": cel,
                "col": PatCol,
                "row": PatRow
            }
            possibleCels.append(cell)
        PatRow += 1
        cel = PatCol + str(PatRow)

    if len(possibleCels) != 0:
        return possibleCels
    #Retentativa mais aberta (as primeiras tres letras da busca
    PatCol = exConsts.PATIENT_COLUMN
    PatRow = exConsts.PAT_ROW
    cel = PatCol + str(PatRow)
    croppedPat = patname.capitalize()[:3]

    while sheet[cel].value is not None:
        if sheet[cel].value.startswith(croppedPat):
            cell = {
                "cll": cel,
                "col": PatCol,
                "row":PatRow
            }
            possibleCels.append(cell)
        PatRow += 1
        cel = PatCol + str(PatRow)

    return possibleCels

def SearchPatientByCell(cell, workbook):
    sheet = workbook.active
    Row = re.search("\d+", cell)[0]
    print(Row)
    nameCell = cell
    PayPerSessionCell = exConsts.PAYPERSESSION_COLUMN + Row
    TotalPaymentCell = exConsts.TOTALPAYMENT_COLUMN + Row
    if sheet[cell].value is not None:
        AllPatientInfo = {
            'name': sheet[nameCell].value,
            'PayPerSession': sheet[PayPerSessionCell].value,
            'TotalPayment': sheet[TotalPaymentCell].value

        }
        return AllPatientInfo


def GetPatientInformation(cell, wb):
    sheet = wb.active
    FirstCell = cell['cll']
    Column = cell['col']
    Row = cell['row']
    PatInfo = []

    ColIn = exConsts.ALPHABET.index(Column)
    NextCol = exConsts.ALPHABET[ColIn+1]
    EndCell = NextCol + str(Row)
    PatRawInfo = sheet[FirstCell:EndCell][0]

    for PRI in PatRawInfo:
        PatInfo.append(PRI.value)

    #Pagamento Total
    Cel = 'O' + str(Row)
    PatInfo.append(sheet[Cel].value)

    PatientDict = {
        'cell': FirstCell,
        'Name': PatInfo[0],
        'PayPerSession': PatInfo[1],
        'TotalPayment': PatInfo[2]
    }

    return PatientDict

def IncrementPayment(Patname, Paidvalue, Month, filepath):
    wb = load_workbook(filepath)
    sheet = wb.active
    PossibleCells = SearchPatientByName(Patname, wb)
    PatientCellComposition = PossibleCells[0]
    row = PatientCellComposition['row']
    col = ''
    for month in exConsts.MONTH_COLS:
        if exConsts.MONTH_COLS[month]['nome'].startswith(Month.capitalize()):
            col = exConsts.MONTH_COLS[month]['col']
            break
    cell = col + str(row)

    sheet[cell].number_format = 'currency'
    sheet[cell].value += int(Paidvalue)
    wb.save(filepath)


