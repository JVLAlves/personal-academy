import re

import exConsts
from openpyxl import load_workbook

import excellentV1.exConsts

"""
def InsertPatient(name, value, workbook):
    sheet = workbook.active
    sheet.insert_rows(exConsts.BEGINNING_ROW)
    Patref = exConsts.PATIENT_COLUMN + str(exConsts.PAT_ROW)
    Valref = exConsts.PAYPERSESSION_COLUMN + str(exConsts.PAT_ROW)
    Totref = exConsts.TOTALPAYMENT_COLUMN + str(exConsts.PAT_ROW)
    sheet[Patref] = name
    sheet[Valref] = value
    sheet[Totref] = '=SUM(C{}:N{})'.format(exConsts.PAT_ROW, exConsts.PAT_ROW)

    workbook.save(exConsts.FILENAME)
"""


def InsertPatient(name, value, filepath):
    workbook = load_workbook(filepath, data_only=True)
    sheet = workbook['entradas2022']
    row = exConsts.BEGINNING_ROW + 1
    cel = exConsts.PATIENT_COLUMN + str(row)
    while sheet[cel].value is not None:
        row += 1
        cel = exConsts.PATIENT_COLUMN + str(row)

    nameCell = cel
    PayPerSessionCell = exConsts.PAYPERSESSION_COLUMN + str(row)
    TotalPaymentCell = exConsts.TOTALPAYMENT_COLUMN + str(row)
    TotalPaymentFormula = '=SUM(C{}:N{})'.format(row, row)
    print(TotalPaymentCell, TotalPaymentFormula)

    sheet[nameCell] = name

    sheet[PayPerSessionCell] = float(value)
    sheet[PayPerSessionCell].number_format = exConsts.FORMAT_CURRENCY_BRL
    sheet[TotalPaymentCell].number_format = exConsts.FORMAT_CURRENCY_BRL
    sheet[TotalPaymentCell] = TotalPaymentFormula

    workbook.save(filepath)


def SearchPatientByName(patname, filename):
    # Variaveis para Uso interno
    workbook = load_workbook(filename, data_only=True)
    sheet = workbook["entradas2022"]  # Planilha
    PatCol = exConsts.PATIENT_COLUMN  # Letra da Coluna do Nome dos Pacientes
    PatRow = exConsts.PAT_ROW  # Número da linha dos Pacientes
    cel = PatCol + str(PatRow)  # Célula determinada com número e letra
    possibleCels = []  # Possiveis células que combinem com a pesquisa.

    # Busca a célula na planilha com o mesmo nome.

    while sheet[cel].value is not None:
        if sheet[cel].value.startswith(patname.capitalize()) or sheet[cel].value.startswith(patname):
            cell = {
                "cll": cel,
                "col": PatCol,
                "row": PatRow
            }
            possibleCels.append(cell)
        PatRow += 1
        cel = PatCol + str(PatRow)

    if len(possibleCels) != 0:
        return possibleCels
    # Retentativa mais aberta (as primeiras tres letras da busca
    PatCol = exConsts.PATIENT_COLUMN
    PatRow = exConsts.PAT_ROW
    cel = PatCol + str(PatRow)
    croppedPat = patname.capitalize()[:3]

    while sheet[cel].value is not None:
        if sheet[cel].value.startswith(croppedPat):
            cell = {
                "cll": cel,
                "col": PatCol,
                "row": PatRow
            }
            possibleCels.append(cell)
        PatRow += 1
        cel = PatCol + str(PatRow)

    return possibleCels


def SearchPatientByCell(cell, filename):
    workbook = load_workbook(filename, data_only=True)
    sheet = workbook['entradas2022']
    RERow = re.search("\d+", cell)
    Row = RERow[0]
    nameCell = cell
    PayPerSessionCell = exConsts.PAYPERSESSION_COLUMN + Row
    TotalPaymentCell = "O" + Row
    if sheet[cell].value is not None:

        AllPatientInfo = {
            'cell': cell,
            'name': sheet[nameCell].value,
            'PayPerSession': sheet[PayPerSessionCell].value,
            'TotalPayment': sheet[TotalPaymentCell].value

        }
        return AllPatientInfo


def GetPatientInformation(cell, filename):
    workbook = load_workbook(filename, data_only=True)
    sheet = workbook.active  # Planilha
    FirstCell = cell['cll']
    Column = cell['col']
    Row = cell['row']
    PatInfo = []

    ColIn = exConsts.ALPHABET.index(Column)
    NextCol = exConsts.ALPHABET[ColIn + 1]
    EndCell = NextCol + str(Row)
    PatRawInfo = sheet[FirstCell:EndCell][0]
    for PRI in PatRawInfo:
        PatInfo.append(PRI.value)

    # Pagamento Total
    Cel = 'O' + str(Row)
    PatInfo.append(sheet[Cel].value)
    print(PatInfo)

    PatientDict = {
        'cell': FirstCell,
        'Name': PatInfo[0],
        'PayPerSession': PatInfo[1],
        'TotalPayment': PatInfo[2]
    }
    return PatientDict


def IncrementPayment(Patname, Paidvalue, Month, filepath):
    print("Will be saving in", filepath)
    wb = load_workbook(filepath, data_only=False)
    sheet = wb['entradas2022']
    PossibleCells = SearchPatientByName(Patname, filepath)
    PatientCellComposition = PossibleCells[0]
    row = PatientCellComposition['row']
    col = ''
    for month in exConsts.MONTH_COLS:
        if isinstance(Month, list):
            Month = Month[0]
        if exConsts.MONTH_COLS[month]['nome'].startswith(Month.capitalize()):
            col = exConsts.MONTH_COLS[month]['col']
            break
    cell = col + str(row)

    if sheet[cell].value is not None:
        sheet[cell].value += float(Paidvalue)
    else:
        sheet[cell].value = float(Paidvalue)
    sheet[cell].number_format = exConsts.FORMAT_CURRENCY_BRL

    self_update(sheet)
    wb.save(filepath)


def DeletePatient(Patname, filepath):
    wb = load_workbook(filepath)
    sheet = wb.active
    PossibleCells = SearchPatientByName(Patname, filepath)
    PatientCellComposition = PossibleCells[0]
    row = PatientCellComposition['row']
    sheet.delete_rows(idx=row, amount=1)
    wb.save(filepath)


def DeleteAll(filepath):
    wb = load_workbook(filepath)
    PossibleCells = SearchPatientByName('', filepath)
    sheet = wb.active
    sheet.delete_rows(idx=excellentV1.exConsts.PAT_ROW, amount=len(PossibleCells))

    wb.save(filepath)

def self_update(sheet):
    row = exConsts.BEGINNING_ROW + 1
    cel = exConsts.PATIENT_COLUMN + str(row)
    TotCel = "O" + str(row)
    while sheet[cel].value is not None:
        TotalPaymentFormula = '=SUM(C{}:N{})'.format(row, row)
        sheet[TotCel].number_format = exConsts.FORMAT_CURRENCY_BRL
        sheet[TotCel] = TotalPaymentFormula
        row += 1
        cel = exConsts.PATIENT_COLUMN + str(row)
        TotCel = "O" + str(row)
